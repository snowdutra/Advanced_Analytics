import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt

# Caminho do arquivo de entrada e saída
input_file_path = r"C:\Users\gustavo.telles\Desktop\Case Itaú\case_estagio_limpo.xlsx"
output_excel_path_updated = r"C:\Users\gustavo.telles\Desktop\Case Itaú\analise_exploraória.xlsx"
# Lendo o arquivo para análise
df = pd.read_excel(input_file_path)

# Converter a coluna 'DataVenda' para datetime
df['DataVenda'] = pd.to_datetime(df['DataVenda'], format='%d/%m/%Y')

# Criar colunas para Ano e Mês
df['Ano'] = df['DataVenda'].dt.year
df['Mês'] = df['DataVenda'].dt.month

# Tendência de vendas por mês e ano
vendas_mensais = df.groupby(['Ano', 'Mês'])['Quantidade'].sum().reset_index()
vendas_mensais['Mês/Ano'] = vendas_mensais['Mês'].astype(str) + '/' + vendas_mensais['Ano'].astype(str)
vendas_mensais_simplificada = vendas_mensais[['Mês/Ano', 'Quantidade']].rename(columns={'Quantidade': 'Quantidade Vendida'})

# Sazonalidade: Soma das quantidades por mês (considerando todos os anos)
sazonalidade = vendas_mensais.groupby('Mês')['Quantidade'].sum()

# Produtos mais vendidos
produtos_vendidos = df.groupby('Produto')['Quantidade'].sum().sort_values(ascending=False).head(10)

# Criar um novo arquivo Excel com todas as análises
wb = Workbook()

# 1. Adicionar dados de Tendência de Vendas (com Mês/Ano combinados)
ws_tendencia = wb.active
ws_tendencia.title = "Tendência de Vendas"

# Adicionar cabeçalhos e dados
tendencia_headers = ['Mês/Ano', 'Quantidade Vendida']
for col_idx, header in enumerate(tendencia_headers, start=1):
    ws_tendencia.cell(row=1, column=col_idx, value=header)
for row_idx, row in enumerate(vendas_mensais_simplificada.itertuples(index=False), start=2):
    for col_idx, value in enumerate(row, start=1):
        ws_tendencia.cell(row=row_idx, column=col_idx, value=value)

# Gerar gráfico de tendência e inserir no Excel
plt.figure(figsize=(12, 6))
for ano in vendas_mensais['Ano'].unique():
    dados_ano = vendas_mensais[vendas_mensais['Ano'] == ano]
    plt.plot(dados_ano['Mês'], dados_ano['Quantidade'], label=f'Ano {ano}')
plt.title('Tendência de Vendas Mensais', fontsize=14)
plt.xlabel('Mês', fontsize=12)
plt.ylabel('Quantidade Vendida', fontsize=12)
plt.legend(title='Ano', loc='upper left', fontsize=10)
plt.grid()
plt.tight_layout(pad=2)  # Ajusta o espaçamento
plt.savefig("tendencia_temp.png", dpi=300)
ws_tendencia.add_image(Image("tendencia_temp.png"), "E1")
plt.close()

# 2. Adicionar dados de Sazonalidade
ws_sazonalidade = wb.create_sheet(title="Sazonalidade")
sazonalidade_headers = ['Mês', 'Quantidade Total Vendida']
ws_sazonalidade.append(sazonalidade_headers)
for mes, quantidade in sazonalidade.items():
    ws_sazonalidade.append([mes, quantidade])

# Gerar gráfico de sazonalidade e inserir no Excel
plt.figure(figsize=(12, 6))
plt.bar(sazonalidade.index, sazonalidade.values, color='orange')
plt.title('Sazonalidade de Vendas (Soma Mensal)', fontsize=14)
plt.xlabel('Mês', fontsize=12)
plt.ylabel('Quantidade Total Vendida', fontsize=12)
plt.xticks(range(1, 13), fontsize=10)
plt.grid(axis='y')
plt.tight_layout(pad=2)  # Ajusta o espaçamento
plt.savefig("sazonalidade_temp.png", dpi=300)
ws_sazonalidade.add_image(Image("sazonalidade_temp.png"), "E1")
plt.close()

# 3. Adicionar dados de Produtos Mais Vendidos
ws_produtos = wb.create_sheet(title="Produtos Mais Vendidos")
produtos_headers = ['Produto', 'Quantidade Vendida']
ws_produtos.append(produtos_headers)
for produto, quantidade in produtos_vendidos.items():
    ws_produtos.append([produto, quantidade])

# Gerar gráfico de produtos mais vendidos e inserir no Excel
plt.figure(figsize=(12, 6))
produtos_vendidos.plot(kind='bar', color='skyblue')
plt.title('Top Produtos Mais Vendidos', fontsize=14)
plt.xlabel('Produto', fontsize=12)
plt.ylabel('Quantidade Vendida', fontsize=12)
plt.xticks(rotation=45, ha='right', fontsize=10)
plt.grid(axis='y')
plt.tight_layout(pad=2)  # Ajusta o espaçamento
plt.savefig("produtos_temp.png", dpi=300)
ws_produtos.add_image(Image("produtos_temp.png"), "E1")
plt.close()

# Salvar o arquivo Excel atualizado
wb.save(output_excel_path_updated)

# Remover imagens temporárias
os.remove("tendencia_temp.png")
os.remove("sazonalidade_temp.png")
os.remove("produtos_temp.png")

print(f"Arquivo Excel salvo com sucesso em: {output_excel_path_updated}")
